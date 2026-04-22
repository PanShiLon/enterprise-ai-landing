"""
第三章：文档解析 - Layer 1 分类解析
支持 .docx / .xlsx / .pdf 三种格式
"""
from pathlib import Path
from dataclasses import dataclass
from typing import List
import re


@dataclass
class RawChunk:
    source: str       # 来源文件名
    content: str      # 原始文本
    chunk_type: str   # faq / manual / table


def parse_document(file_path: str) -> List[RawChunk]:
    path = Path(file_path)
    ext = path.suffix.lower()

    if ext == ".docx":
        return _parse_word(path)
    elif ext in (".xlsx", ".xls"):
        return _parse_excel(path)
    elif ext == ".pdf":
        return _parse_pdf(path)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")


def _parse_word(path: Path) -> List[RawChunk]:
    from docx import Document
    doc = Document(str(path))
    chunks = []
    buffer = []

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        # 按标题分块
        if para.style.name.startswith("Heading"):
            if buffer:
                chunks.append(RawChunk(
                    source=path.name,
                    content="\n".join(buffer),
                    chunk_type="manual"
                ))
                buffer = []
        buffer.append(text)

    if buffer:
        chunks.append(RawChunk(
            source=path.name,
            content="\n".join(buffer),
            chunk_type="manual"
        ))

    return chunks


def _get_merged_value(ws, row, col):
    """读取合并单元格的值（取左上角主单元格）"""
    for merged_range in ws.merged_cells.ranges:
        if (row, col) in merged_range:
            return ws.cell(merged_range.min_row, merged_range.min_col).value
    return ws.cell(row, col).value


def _parse_excel(path: Path) -> List[RawChunk]:
    # 旧版 .xls 先转换
    if path.suffix.lower() == ".xls":
        import xlrd
        import openpyxl
        xls = xlrd.open_workbook(str(path))
        wb = openpyxl.Workbook()
        for sheet_name in xls.sheet_names():
            xls_sheet = xls.sheet_by_name(sheet_name)
            ws = wb.create_sheet(title=sheet_name)
            for row in range(xls_sheet.nrows):
                for col in range(xls_sheet.ncols):
                    ws.cell(row=row + 1, column=col + 1,
                            value=xls_sheet.cell_value(row, col))
        import tempfile, os
        tmp = tempfile.mktemp(suffix=".xlsx")
        wb.save(tmp)
        path = Path(tmp)

    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True)
    chunks = []

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows())
        if not rows:
            continue

        # 自动探测表头行：有效值密度 > 50%
        header_row = 0
        for i, row in enumerate(rows[:5]):
            non_empty = sum(1 for c in row if c.value is not None)
            if non_empty > len(row) * 0.5:
                header_row = i
                break

        headers = [str(rows[header_row][c].value or "").strip()
                   for c in range(len(rows[header_row]))]

        for row in rows[header_row + 1:]:
            values = [_get_merged_value(sheet, row[c].row, c + 1)
                      for c in range(len(headers))]
            row_dict = dict(zip(headers, values))
            # 过滤空行
            if not any(v for v in values if v):
                continue
            content = " | ".join(
                f"{k}: {v}" for k, v in row_dict.items() if v
            )
            chunks.append(RawChunk(
                source=path.name,
                content=content,
                chunk_type="table"
            ))

    return chunks


def _parse_pdf(path: Path) -> List[RawChunk]:
    try:
        import pdfplumber
        chunks = []
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text and len(text.strip()) > 20:
                    chunks.append(RawChunk(
                        source=path.name,
                        content=text.strip(),
                        chunk_type="manual"
                    ))
        return chunks
    except Exception as e:
        print(f"⚠️  PDF解析失败: {path.name} - {e}")
        print("   提示：扫描件/特殊字体PDF需要人工复制粘贴")
        return []


FAQ_PATTERNS = [
    r"Q[：:]\s*(.+?)\nA[：:]\s*(.+?)(?=\nQ|$)",
    r"问题[：:]\s*(.+?)\n答[：:]\s*(.+?)(?=\n问题|$)",
    r"【问】(.+?)【答】(.+?)(?=【问】|$)",
]


def extract_faq(text: str) -> List[tuple]:
    """从FAQ型文本中提取 (question, answer) 对"""
    for pattern in FAQ_PATTERNS:
        pairs = re.findall(pattern, text, re.DOTALL)
        if pairs:
            return [(q.strip(), a.strip()) for q, a in pairs]
    return []
